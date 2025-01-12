import openpyxl
from io import BytesIO
import streamlit as st

st.title("Language and Cognition Lab🧠📊")
st.image('https://i.ibb.co/L0hHCFZ/Screenshot-2023-02-22-at-12-57-43-PM.png')
st.markdown("Automated Excel Editor")
excel_file = st.file_uploader('Upload your excel file')

def same_choice(p_column, j_column):
    result = []
    for p, j in zip(p_column, j_column):
        if p.value == j.value:
            result.append(1)
        elif p.value == "don't know":
            result.append(0.5)
        else:
            result.append(0)
    return result

def choice(m_column, l_column, k_column):
    result = []
    for m, l, k in zip(m_column, l_column, k_column):
        if m.value == 'j':
            result.append(l.value)
        elif m.value == 'f':
            result.append(k.value)
        elif m.value == 'd':
            result.append('don\'t know')
        else:
            result.append(None)
    return result

def get_belief_type(column):
    results = []
    for cell in column:
        string = str(cell.value)
        results.append(string[-1])
    return results

def edit_excel():

  wb_old = openpyxl.load_workbook(BytesIO(excel_file.read()))
  ws_old = wb_old.active #specify the ws_old name to select other than the active ws_old
  number = int(ws_old['A12'].value.replace('Subject Number: ', ''))
  if ws_old.max_row > 236:
    st.warning("The spreadsheet provided is missing the number of data required. You should have samples 2 to 25. Every sample should have exactly 9 trials/rows.", icon="⚠️")

  else:
    print("Maximum rows before removing : ",ws_old.max_row)
    i = 2 # sample numbers start at 2\
    def count_trials(ws_old, sample_number):
        """Counts how many times a sample appears in column A of ws_old."""
        count = 0
        for i in range(1, ws_old.max_row + 1):  # Iterate through all rows in column A
            cell_value = ws_old.cell(row=i, column=1).value  # Access the value directly in column A (1st column)
            # print(f"Row {i} | Cell Value: {cell_value}")
            
            if cell_value is None or cell_value == '':  # Skip empty or None values
                continue
            
            cell_value = str(cell_value).strip()  # Convert value to string and strip spaces
            # print(f"Processed Value: {cell_value}")

            if cell_value == str(sample_number):  # Compare as strings
                count += 1

        return count

    j = count_trials(ws_old, i)
    ws_old.insert_cols(1, 2)
    ws_old.insert_cols(idx=4)
    ws_old.delete_rows(0,17)

    ws_old['A1'] = "Subject Number"
    ws_old['B1'] = 'list'
    ws_old['D1'] = 'null'
    for row in ws_old.iter_rows():
        values = [cell.value for cell in row]
        values[4], values[10] = values[10], values[4]
        for i, cell in enumerate(row):
            cell.value = values[i]
            
    path1 = 'edited_.xlsx'
    #   number = subject_number()
    wb = openpyxl.Workbook()
    column_values = ['Subject Number', 'list', 'trial', 'null', 'condition', 'time',
                    'Relationship',
                    'ControlQ1 Copy 2', 'ControlQ1 Copy - 2 - 2',
                    'FirstMoozleProp Copy 13', 'SecondMoozleProp Copy 13',
                    'SecondMoozleProp2 Copy 13','ChoiceResponse Copy 2',
                    'ControlQ2 Copy 2', 'ControlQ2 Copy-2 - 2', 'Choice','SameChoice',
                    'BeliefType', 'AgeGroup']
    ws = wb.active

    for index, value in enumerate(column_values, start=1):
        cell = ws.cell(row=1, column=index)
        cell.value = value
        wb.save(path1)

    # ws_old.delete_rows(0,17)

    print('j: ',j)
    while i <= 25:
        for cell in ws_old['C']:
            if cell.value == i:
                try:
                    if i == 2:

                    #Choice
                        ws.cell(row=i, column=16).value = choice(ws['M'], ws['L'], ws['K'])[i-2]

                        # Subject Number
                        ws.cell(row=i, column=1).value = number

                        # Trial
                        string = ws_old.cell(row=i, column=3).value
                        ws.cell(row=i, column=3).value = string

                        # Condition
                        string = ws_old.cell(row=i, column=6).value
                        ws.cell(row=i, column=5).value = string.replace('.PICT @ :Pictures:', '')

                        # Relationship
                        string = ws_old.cell(row=i, column=5).value
                        ws.cell(row=i, column=7).value = string.replace('.PICT @ :Pictures:', '')
                        
                        #ControlQ1 Copy 2 the 2nd row for the trial in keys between []
                        string = ws_old.cell(row=i+1, column=14).value
                        ws.cell(row=i, column=8).value = string.replace('[', '').replace(']','')

                        # ControlQ1 Copy - 2 - 2the 3rd  row for the trial in keys
                        string = ws_old.cell(row=i+2, column=14).value
                        ws.cell(row=i, column=9).value = string.replace('[', '').replace(']','')
            
                        #FirstMoozleProp Copy 13 the 4th row for the trial in response label
                        string = ws_old.cell(row=i+3, column=5).value
                        ws.cell(row=i, column=10).value = string.replace('.PICT @ :Pictures:', '')

                        # SecondMoozleProp Copy 13 the 5th row for the trial in response label
                        string = ws_old.cell(row=i+4, column=5).value
                        ws.cell(row=i, column=11).value = string.replace('.PICT @ :Pictures:', '')


                        #SecondMoozleProp2 Copy 13he 6th row for the trial in response
                        string = ws_old.cell(row=i+5, column=5).value
                        ws.cell(row=i, column=12).value = string.replace('.PICT @ :Pictures:', '')

                        # time the 6th row for that trial in Time
                        string = ws_old.cell(row=i+6, column=12).value
                        ws.cell(row=i, column=6).value = string

                        #ChoiceResponse Copy 2 the 7th row for the trial in keys between []
                        string = ws_old.cell(row=i+6, column=14).value
                        ws.cell(row=i, column=13).value = string.replace('[', '').replace(']','')


                        #ControlQ2 Copy 2 the 8th row for the trial in keys between
                        string = ws_old.cell(row=i+7, column=14).value
                        ws.cell(row=i, column=14).value = string.replace('[', '').replace(']','')
                        
                        #ControlQ2 Copy-2 - 2 the 9th row for the trial in keys between
                        string = ws_old.cell(row=i+8, column=14).value
                        ws.cell(row=i, column=15).value = string.replace('[', '').replace(']','')

                        i += 1
                    else:
                        # Subject Number
                        ws.cell(row=i, column=1).value = number              

                        # Trial
                        string = ws_old.cell(row=j, column=3).value
                        ws.cell(row=i, column=3).value = string

                        # Condition
                        string = ws_old.cell(row=j, column=6).value
                        ws.cell(row=i, column=5).value = string.replace('.PICT @ :Pictures:', '')

                        # Relationship
                        string = ws_old.cell(row=j, column=5).value
                        ws.cell(row=i, column=7).value = string.replace('.PICT @ :Pictures:', '')

                        #ControlQ1 Copy 2 the 2nd row for the trial in keys between []
                        string = ws_old.cell(row=j+1, column=14).value
                        ws.cell(row=i, column=8).value = string.replace('[', '').replace(']','')

            
                        #ControlQ1 Copy - 2 - 2the 3rd  row for the trial in keys
                        string = ws_old.cell(row=j+2, column=14).value
                        ws.cell(row=i, column=9).value = string.replace('[', '').replace(']','')
            

                        #FirstMoozleProp Copy 13 the 4th row for the trial in response label
                        string = ws_old.cell(row=j+3, column=5).value
                        ws.cell(row=i, column=10).value = string.replace('.PICT @ :Pictures:', '')

                        # SecondMoozleProp Copy 13 the 5th row for the trial in response label
                        string = ws_old.cell(row=j+4, column=5).value
                        ws.cell(row=i, column=11).value = string.replace('.PICT @ :Pictures:', '')

                        #SecondMoozleProp2 Copy 13he 6th row for the trial in response
                        string = ws_old.cell(row=j+5, column=5).value
                        ws.cell(row=i, column=12).value = string.replace('.PICT @ :Pictures:', '')


                        # time the 6th row for that trial in Time
                        string = ws_old.cell(row=j+6, column=12).value
                        ws.cell(row=i, column=6).value = string

                        #ChoiceResponse Copy 2 the 7th row for the trial in keys between []
                        string = ws_old.cell(row=j+6, column=14).value
                        ws.cell(row=i, column=13).value = string.replace('[', '').replace(']','')

                        #ControlQ2 Copy 2 the 8th row for the trial in keys between
                        string = ws_old.cell(row=j+7, column=14).value
                        ws.cell(row=i, column=14).value = string.replace('[', '').replace(']','')

                        #ControlQ2 Copy-2 - 2 the 9th row for the trial in keys between
                        string = ws_old.cell(row=j+8, column=14).value
                        ws.cell(row=i, column=15).value = string.replace('[', '').replace(']','')

                        i += 1
                        j += 9 
                        print(i,j)
                except:
                    i += 1
                    j += 9  
                    print(i,j)
                    continue

    for i in range(24):
        # print(choice(ws['M'], ws['L'], ws['K'])[i+1])
        ws.cell(row=i+2, column=16).value = choice(ws['M'], ws['L'], ws['K'])[i+1]
        ws.cell(row=i+2, column=17).value = same_choice(ws['P'], ws['J'])[i+1]
        ws.cell(row=i+2, column=18).value = get_belief_type(ws['E'])[i+1]

    path1 = 'edited_.xlsx'
    wb.save(path1)
    st.download_button(
            label="Download Updated Excel Workbook",
            data=open(path1, 'rb').read(),
            file_name="workbook.xlsx",
            mime="xlsx"
        )


        
st.write('''
Please make sure to have the header of the original excel file be between rows 1-17.
Also make sure to have the Subject Number on cell A12.\nEnjoy!!''')
      
edit = st.button('Click me')
if edit:
    edit_excel()